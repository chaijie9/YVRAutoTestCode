import os
import re
import subprocess
import time
import openpyxl
from dingtalkchatbot.chatbot import DingtalkChatbot
# from slamawake import adb_command check_adb_device


webhook = 'https://oapi.dingtalk.com/robot/send?access_token' \
          '=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622 '
ding = DingtalkChatbot(webhook)
Excel_File = r"./reboot_D3_RC_D3.0.2.45+0424.xlsx"

if not os.path.exists(Excel_File):
    wb = openpyxl.Workbook()
    wb.create_sheet(title=str("reboot_MSG"), index=0)
    wb.save(Excel_File)
workbook1 = openpyxl.load_workbook(Excel_File)
sh = workbook1["reboot_MSG"]
ItemIndex = ["测试次数", "初始化是否成功", "mapping重定位成功", "crash" ]
count = 1
for Item in ItemIndex:
    sh.cell(1, count, value=Item)
    count = count + 1
workbook1.save(Excel_File)
test_count = 1
error = 0
# sn = "192.168.8.107:5555"
sn = "D3HDXD2D4363000200"

# connect = "adb connect 192.168.8.116:5555"
connect = f"adb connect {sn}"
def check_adb_device():
    try:
        result = subprocess.run(['adb', 'devices'], capture_output=True, text=True, timeout=5)
        output = result.stdout
        if f'{sn}      device' in output:
            return False
        else:
            return True
    except subprocess.TimeoutExpired:
        return False
    except Exception as e:
        return False

def adb_command(command):
    try:
        result = subprocess.run(command, capture_output=True, check=True, text=True, timeout=30, encoding="utf-8",shell=True)
        msg = str(result.stdout).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        return msg
    except subprocess.CalledProcessError as e:
        ding.send_text(msg=time.strftime("%H:%M:%S ") + f' {e} 无结果', is_at_all=False)
        return f"未获取到当前结果 {e}"
    except subprocess.TimeoutExpired as e:
        return f"{command} 命令运行超时\n{e}"

while True:
    sh.cell(1 + test_count, 1, value= test_count)
    while True:
        # if not check_adb_device():
        #     subprocess.call(connect,shell=True)
        #     print(check_adb_device())
        #     time.sleep(2)
        # else:
        start_modeC = f"adb -s {sn} shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
        subprocess.call(start_modeC, shell=True, timeout=30)
        time.sleep(2)
        root = f"adb -s {sn} shell setprop service.dev.mode 1"
        print(root)
        subprocess.call(root, shell=True, timeout=60)
        time.sleep(3)

        # 检查slam初始化#
        init_command = f'adb -s {sn} logcat -d | findstr "delivered."'
        p_obj = subprocess.Popen(
            args=init_command,
            stdin=None, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, shell=True)
        for line in p_obj.stdout:
            msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
                .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
        if msg.find("6DOF is ready!!!!!!!!!") >= 0:
            # result = re.findall(r"6DOF is ready!!!!!!!!!")
            sh.cell(1 + test_count, 2, value="PASS")
            print("初始化成功")
        else:
            now_time = time.strftime("%H_%M_%S")
            subprocess.call(f"adb -s {sn} logcat -d > ./LOG/6DofInit_{now_time}_{test_count}次.txt", shell=True)
            sh.cell(1 + test_count, 2, value="FAIL")
            print("初始化失败")

        # 检查mapping重定位
        recenter_command = f'adb -s {sn} logcat -d | findstr "Merge"'
        p_obj = subprocess.Popen(
            args=recenter_command,
            stdin=None, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, shell=True)
        for line in p_obj.stdout:
            msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
                .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
        time.sleep(2)
        msg = adb_command(recenter_command)
        if msg.find("mapping Merge global map") >= 0:
            sh.cell(1 + test_count, 3, value="PASS")
            print("mapping 重定位成功")
        else:
            sh.cell(1 + test_count, 3, value="FAIL")
            print("mapping 重定位失败")
            now_time = time.strftime("%H_%M_%S")

        # 检查有无crash
        result_tomb = adb_command(f"adb -s {sn} shell ls data/tombstones/")
        result_anr = adb_command(f"adb -s {sn} shell ls data/anr/")
        if result_tomb.find("_") > 0 or result_anr.find("_") > 0:
            now_time = time.strftime("%H_%M_%S")
            subprocess.call(f"adb -s {sn} logcat -d > ./LOG/crash_{now_time}_{test_count}.txt", shell=True)
            subprocess.call(f"adb -s {sn} pull /data/tombstones  ./tomb/{now_time}/")
            time.sleep(2)
            subprocess.call(f"adb -s {sn} shell rm -rf /data/tombstones/*", shell=True)
            subprocess.call(f"adb -s {sn} shell rm -rf /data/anr/*", shell=True)
            time.sleep(2)
            ding.send_text(sn + time.strftime("%H:%M:%S") + f' 第 {test_count} 次重启测试后crash', is_at_all=False)
            sh.cell(1 + test_count, 4, value="YES")
        else:
            sh.cell(1 + test_count, 4, value="NO")
            print(f"第 {test_count} 次，未复现crash")
        subprocess.call(f"adb -s {sn} reboot",shell=True)
        print("reboot")
        time.sleep(65)
        test_count += 1
        workbook1.save(Excel_File)
    workbook1.save(Excel_File)


