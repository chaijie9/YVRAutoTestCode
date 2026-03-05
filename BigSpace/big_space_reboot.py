import os
import re
import subprocess
import time
import openpyxl
from dingtalkchatbot.chatbot import DingtalkChatbot
# from slamawake import adb_command

def check_adb_device():
    try:
        result = subprocess.run(['adb', 'devices'], capture_output=True, text=True, timeout=5)
        output = result.stdout
        if '192.168.8.106:5555	device' in output:
            return False
        else:
            return True
    except subprocess.TimeoutExpired:
        return False
    except Exception as e:
        return False

def adb_command(command):
    try:
        result = subprocess.run(command, capture_output=True, check=True, text=True, timeout=30, encoding="utf-8",shell=True)
        msg = str(result.stdout).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        return msg
    except subprocess.CalledProcessError as e:
        ding.send_text(msg=time.strftime("%H:%M:%S ") + f' {e} 无结果', is_at_all=False)
        return f"未获取到当前结果 {e}"
    except subprocess.TimeoutExpired as e:
        return f"{command} 命令运行超时\n{e}"



webhook = 'https://oapi.dingtalk.com/robot/send?access_token' \
          '=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622 '
ding = DingtalkChatbot(webhook)
Excel_File = r"./reboot_D2_ENT_2.1.1.126+0.4.88.xlsx"

if not os.path.exists(Excel_File):
    wb = openpyxl.Workbook()
    wb.create_sheet(title=str("reboot_MSG"), index=0)
    wb.save(Excel_File)
workbook1 = openpyxl.load_workbook(Excel_File)
sh = workbook1["reboot_MSG"]
ItemIndex = ["测试次数", "SLAM初始化", "crash" ]
count = 1
for Item in ItemIndex:
    sh.cell(1, count, value=Item)
    count = count + 1
workbook1.save(Excel_File)
test_count = 1
error = 0
sn = "192.168.8.106:5555"
# connect = "adb connect 192.168.8.116:5555"
connect = "adb connect 192.168.8.106:5555"
while True:
    sh.cell(1 + test_count, 1, value= test_count)
    while True:
        # if check_adb_device():
        #     subprocess.call(connect)
        #     time.sleep(5)
        # else:
        time.sleep(3)
        start_modeC = f"adb -s {sn} shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
        subprocess.call(start_modeC, shell=True, timeout=30)
        time.sleep(5)
        root = f"adb -s {sn} shell setprop service.dev.mode 1"
        print(root)
        subprocess.call(root, shell=True, timeout=60)
        time.sleep(3)
            #
            #
            # # 地图文件名称修改并导入
            # file_name = os.listdir("./_Maps/.")[0]
            # parts = file_name.split('-')
            # last_part = parts[-1]
            # new_last_part = str(int(last_part) + 1)
            # parts[-1] = new_last_part
            # new_folder_name = '-'.join(parts)
            # os.rename(f"./_Maps/{file_name}", f"./_Maps/{new_folder_name}")
            # print(f"更换地图名称成功{new_folder_name}")
            # time.sleep(2)
            # adb_command("adb push _Maps/. /sdcard/Maps/")
            # sh.cell(1 + test_count, 2, value="PASS")
            # print("==========导入成功============")

            # 等待设备重启；
            # subprocess.call("adb remount")

        # 检查有无crash
        result_tomb = adb_command(f"adb -s {sn} shell ls data/tombstones/")
        result_anr = adb_command(f"adb -s {sn} shell ls data/anr/")
        if result_tomb.find("_") > 0 or result_anr.find("_") > 0:
            now_time = time.strftime("%H_%M_%S")
            subprocess.call(f"adb -s {sn} logcat -d > ./LOG/crash_{now_time}_{test_count}.txt", shell=True)
            subprocess.call(f"adb -s {sn} pull /data/tombstones  ./tomb/{now_time}/")
            time.sleep(2)
            subprocess.call(f"adb -s {sn} shell rm -rf /data/tombstones/*", shell=True)
            subprocess.call(f"adb -s {sn} shell rm -rf /data/anr/*", shell=True)
            time.sleep(2)
            ding.send_text(sn + time.strftime("%H:%M:%S") + f' 第 {test_count} 次重启测试后crash', is_at_all=False)
            sh.cell(1 + test_count, 4, value="YES")
        else:
            sh.cell(1 + test_count, 4, value="NO")
            print(f"第 {test_count} 次，未复现crash")

        #检查slam初始化
        init_command = f'adb -s {sn} logcat -d | findstr "delivered."'
        p_obj = subprocess.Popen(
            args=init_command,
            stdin=None, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, shell=True)
        for line in p_obj.stdout:
            msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
                .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
        if msg.find("6DOF is ready!!!!!!!!!") >= 0:
            # result = re.findall(r"6DOF is ready!!!!!!!!!")
            sh.cell(1 + test_count, 2, value="PASS")
            print("初始化成功")
        else:
            now_time = time.strftime("%H_%M_%S")
            subprocess.call(f"adb -s {sn} logcat -d > ./LOG/6DofInit_{now_time}_{test_count}次.txt", shell=True)
            sh.cell(1 + test_count, 2, value="FAIL")
            print("初始化失败")
            time.sleep(2)


            # 检查mapping重定位
        recenter_command = f'adb -s {sn} logcat -d | findstr "large_space_map_recognized"'
        p_obj = subprocess.Popen(
            args=recenter_command,
            stdin=None, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, shell=True)
        for line in p_obj.stdout:
            msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
                .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
        if msg.find("large_space_map_recognized setRecenterPose") >= 0:
            sh.cell(1 + test_count, 3, value="PASS")
            print("mapping 重定位成功")
        else:
            sh.cell(1 + test_count, 3, value="FAIL")
            print("mapping 重定位失败")
            time.sleep(2)
            subprocess.call(f"adb -s {sn}  logcat -d > ./LOG/{now_time}_{test_count}次.txt", shell=True)
            time.sleep(3)
        # recenter_command = f'adb -s {sn} logcat -d | findstr "FindContours_dsp"'
        # p_obj = subprocess.Popen(
        #     args=recenter_command,
        #     stdin=None, stdout=subprocess.PIPE,
        #     stderr=subprocess.PIPE, shell=True)
        # for line in p_obj.stdout:
        #     msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
        #         .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
        # if msg.find("FindContours_dsp") <= 0:
        #     sh.cell(1 + test_count, 3, value="PASS")
        #     print("0609未复现手柄问题")
        # else:
        #     sh.cell(1 + test_count, 3, value="FAIL")
        #     print("0609疑似复现手柄问题")
        #     now_time = time.strftime("%H_%M_%S")
        #     subprocess.call(f"adb logcat -d > ./LOG/{now_time}_{test_count}次.txt", shell=True)
        #     time.sleep(3)
        #     ding.send_text(time.strftime("%H:%M:%S") + f' 第 {test_count} 0609疑似复现手柄问题', is_at_all=False)


            workbook1.save(Excel_File)
            # 清空地图数据
            # subprocess.call("adb shell rm -rf /sdcard/Maps/.")
            # time.sleep(2)
            # subprocess.call("adb shell rm -rf /data/misc/trackingservice/algs_bs/.")
            # time.sleep(2)
            # subprocess.call("adb shell setprop persist.yvr.large_space_enable 0")
            # print("set 0")
            # time.sleep(2)
        subprocess.call(f"adb  -s {sn} reboot")
        print("reboot")
        time.sleep(65)
        test_count += 1
    workbook1.save(Excel_File)


