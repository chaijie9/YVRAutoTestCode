import os
import re
import subprocess
import time

import openpyxl
from dingtalkchatbot.chatbot import DingtalkChatbot
# from slamawake import adb_command



def adb_command(command):
    try:
        result = subprocess.run(command, capture_output=True, check=True, text=True, timeout=30, encoding="utf-8",shell=True)
        msg = str(result.stdout).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        return msg
    except subprocess.CalledProcessError as e:
        ding.send_text(msg=time.strftime("%H:%M:%S ") + f' {e} 无结果', is_at_all=False)
        return f"未获取到当前结果 {e}"
    except subprocess.TimeoutExpired as e:
        return f"{command} 命令运行超时\n{e}"


now_time = time.strftime("%H_%M_%S")
webhook = 'https://oapi.dingtalk.com/robot/send?access_token' \
          '=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622 '
ding = DingtalkChatbot(webhook)
# Excel_File = r"./reboot_RC_mapping_debug_crash_250106.xlsx"
# if not os.path.exists(Excel_File):
#     wb = openpyxl.Workbook()
#     wb.create_sheet(title=str("Slam_Log_MSG"), index=0)
#     wb.save(Excel_File)
# workbook1 = openpyxl.load_workbook(Excel_File)
# sh = workbook1["Slam_Log_MSG"]
# ItemIndex = ["测试次数", "初始化是否成功", "mapping重定位成功", "crash" ]
# count = 1
# for Item in ItemIndex:
#     sh.cell(1, count, value=Item)
#     count = count + 1
# workbook1.save(Excel_File)
test_count = 1
error = 0
while True:
    # subprocess.call("adb connect 192.168.8.116:5555")
    # time.sleep(3)
    # update_sleep_time = "adb -s D3HDXD2D4363000102 shell logcat -c"
    # subprocess.call(update_sleep_time)
    subprocess.call("adb reboot ")
    # sh.cell(1 + test_count, 1, value=test_count)
    # subprocess.call("adb -s D3HDXD2D4363000102  shell input keyevent 223") # off
    time.sleep(65)
    # subprocess.call("adb -s D3HDXD2D4363000186 shell am start -n com.oculus.sdk.yvrmesh/com.oculus.NativeActivity")
    # print("开MESH")
    start_modeC = "adb  shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
    print(start_modeC)
    subprocess.call(start_modeC, shell=True, timeout=15)
    time.sleep(3)
    root = "adb shell setprop service.dev.mode 1"
    print(root)
    subprocess.call(root, shell=True, timeout=10)
    time.sleep(5)
    # subprocess.call("adb -s D3HDXD2D4363000102 shell input keyevent 224") # on
    print("start")
    time.sleep(8)
    init_command = 'adb  logcat -d | findstr "delivered."'
    p_obj = subprocess.Popen(
        args=init_command,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
    if msg.find("Low frequency vio state is delivered. 6DOF is ready!!!!!!") >= 0:
        # result = re.findall(r"6DOF is ready!!!!!!!!!")
        # sh.cell(1 + test_count, 2, value="PASS")
        print("初始化成功")
    else:
        # sh.cell(1 + test_count, 2, value="FAIL")
        print("初始化失败")
        time.sleep(2)

    # recenter_command = 'adb -s D3HDXD2D4363000186 logcat -d | findstr "large_space_map_recognized"'
    # p_obj = subprocess.Popen(
    #     args=recenter_command,
    #     stdin=None, stdout=subprocess.PIPE,
    #     stderr=subprocess.PIPE, shell=True)
    # for line in p_obj.stdout:
    #     msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
    #         .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
    # if msg.find("large_space_map_recognized setRecenterPose") >= 0:
    #     sh.cell(1 + test_count, 3, value="PASS")
    #     print("mapping 重定位成功")
    # else:
    #     sh.cell(1 + test_count, 3, value="FAIL")
    #     print("mapping 重定位失败")
    #     time.sleep(2)
    #     subprocess.call(f"adb -s D3HDXD2D4363000186  logcat -d > ./LOG/{now_time}_{test_count}次.txt", shell=True)
    # time.sleep(3)
    # subprocess.call("adb shell dumpsys vrruntimeservice_native --do_recenter")

    # recenter_command = 'adb logcat -d | findstr "Merge"'
    # p_obj = subprocess.Popen(
    #     args=recenter_command,
    #     stdin=None, stdout=subprocess.PIPE,
    #     stderr=subprocess.PIPE, shell=True)
    # for line in p_obj.stdout:
    #     msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
    #         .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
    # time.sleep(5)
    # msg = adb_command(recenter_command)
    # if msg.find("mapping Merge global map") >= 0:
    #     # sh.cell(1 + test_count, 3, value="PASS")
    #     print("mapping 重定位成功")
    # else:
    #     # sh.cell(1 + test_count, 3, value="FAIL")
    #     print("mapping 重定位失败")
    #     time.sleep(2)
    #     now_time = time.strftime("%H_%M_%S")
        # subprocess.call(f"adb -s 192.168.8.115:5555 logcat -d > ./LOG/{now_time}_{test_count}.txt", shell=True)
        # subprocess.call(f"adb -s D3HDXD2D4363000186 logcat -d > ./LOG/{now_time}_{test_count}.txt", shell=True)

    print(f"第", test_count, "次")
    result_tomb = adb_command("adb  shell ls data/tombstones/")
    result_anr = adb_command("adb  shell ls data/anr/")
    print(result_tomb, result_anr)


    if result_tomb.find("_") > 0 or result_anr.find("_") > 0:
        now_time = time.strftime("%H_%M_%S")
        subprocess.call(f"adb  logcat -d > ./LOG/crash_{now_time}_{test_count}次.txt", shell=True)
        subprocess.call(f"adb  pull /data/tombstones  ./tomb/{now_time}")
        time.sleep(2)
        subprocess.call("adb shell rm -rf /data/tombstones/*", shell=True)
        subprocess.call("adb shell rm -rf /data/anr/*", shell=True)
        time.sleep(2)
        ding.send_text(time.strftime("%H:%M:%S") + f' 第 {test_count} 次重启后crash', is_at_all=False)
    else:
        # sh.cell(1 + test_count, 4, value="NO")
        print(f"第 {test_count} 次，当前无crash")
        test_count += 1
    # workbook1.save(Excel_File)
# workbook1.save(Excel_File)