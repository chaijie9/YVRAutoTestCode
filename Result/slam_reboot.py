
# start_modeC = "adb shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
# print(start_modeC)
# subprocess.call(start_modeC, shell=True, timeout=5)
# print("start ")
# root = "adb wait-for-device shell setprop service.dev.mode 1"
# print(root)
# subprocess.call(root, shell=True, timeout=5)
# print("==============Root==============")
# time.sleep(3)
# # 关闭静置检测
# check = "adb shell getprop persist.sys.movecheck.enable"
# off_check = "adb shell setprop persist.sys.movecheck.enable 0"
# if subprocess.call(check, shell=True, timeout=15) is True:
import os
import re
import subprocess
import time
import openpyxl
from dingtalkchatbot.chatbot import DingtalkChatbot
Excel_File = "reboot_hand.xlsx"
error = 0
# 当前路径创建excel保存测试结果
if not os.path.exists(Excel_File):
    wb = openpyxl.Workbook()
    wb.create_sheet(title=str("Slam_Log_MSG"), index=0)
    wb.save(Excel_File)
workbook1 = openpyxl.load_workbook(Excel_File)
sh = workbook1["Slam_Log_MSG"]
ItemIndex = ["测试次数", "Slam初始化时间", "边界数据", "当前是否在边界内"]
trace = 0
count = 1
for Item in ItemIndex:
    sh.cell(1, count, value=Item)
    count = count + 1
workbook1.save(Excel_File)
now_time = time.strftime("%H:%M:%S")
webhook = 'https://oapi.dingtalk.com/robot/send?access_token' \
          '=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622 '
ding = DingtalkChatbot(webhook)
test_count = 0
while test_count <= 1000:
    connect = "adb connect 172.16.45.216:5555"
    # subprocess.call(connect, shell=True, timeout=5)
    start_modeC = "adb shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
    print(start_modeC)
    subprocess.call(start_modeC, shell=True, timeout=15)
    time.sleep(3)
    root = "adb wait-for-device shell setprop service.dev.mode 1"
    print(root)
    subprocess.call(root, shell=True, timeout=5)
    print("start")
    time.sleep(4)
    # subprocess.call(connect, shell=True, timeout=5 )
    sh.cell(1 + test_count, 1, value=test_count)
    subprocess.call("adb root", shell=True, timeout=5)
   #  check slam init time
    command = 'adb logcat -d | grep "EnterVRMode cost"'
    p_obj = subprocess.Popen(
            args=command,
            stdin=None, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find(r"EnterVRMode cost") >= 0:
            result_time = re.findall(r"EnterVRMode cost (.*?) s", str(msg))
            result_time = result_time[0]
            sh.cell(1 + test_count, 2, value=result_time)
            print(f"初始化时间：{result_time}")
        else:
            error += 1
            sh.cell(1 + test_count, 2, value="null")
    # check boundary data
    result = subprocess.check_output("adb shell ls data/misc/gds/boundary/"  ,  shell=True)
    if str(result).find(".") <= 0:
        ding.send_text(msg=f"第{test_count} 次重启异常请排查" + time.strftime("%H:%M:%S"), is_at_all=False)
        error += 1
    else:
        print(f"{test_count} -- pass")
        sh.cell(1 + test_count, 3, value="pass")
    # check in boundary
    in_boundary_command = 'adb logcat -d| findstr "notifyMotion_type===2"'
    p_obj = subprocess.Popen(
        args=in_boundary_command,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", "")
        if msg.find(r"notifyMotion_type===2") >= 0:
            print("in_boundary")
            sh.cell(1 + test_count, 4, value="pass")
        else:
            error += 1
            sh.cell(1 + test_count, 4, value=f"{now_time} error")
            print("fail")
    workbook1.save(Excel_File)
    if error == 0:
        subprocess.call("adb reboot", shell=True, timeout=5)
        print(f"第{test_count} 次重启")
        time.sleep(35)
        test_count += 1
    else:
        break
# 3-3
# 4-4
# 5-5
#