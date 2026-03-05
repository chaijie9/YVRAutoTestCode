import os
import re
import subprocess
import time
import openpyxl
from dingtalkchatbot.chatbot import DingtalkChatbot
from slamawake import adb_command




webhook = 'https://oapi.dingtalk.com/robot/send?access_token' \
          '=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622 '
ding = DingtalkChatbot(webhook)
Excel_File = r"./reboot_D3_RC_D3.0.1.44+resAlg"
if not os.path.exists(Excel_File):
    wb = openpyxl.Workbook()
    wb.create_sheet(title=str("import_MSG"), index=0)
    wb.save(Excel_File)
workbook1 = openpyxl.load_workbook(Excel_File)
sh = workbook1["Reboot_MSG"]
ItemIndex = ["测试次数", "SLAM初始化", "重定位", "crash" ]
count = 1
for Item in ItemIndex:
    sh.cell(1, count, value=Item)
    count = count + 1
workbook1.save(Excel_File)


test_count = 1
error = 0
while True:
    sh.cell(1 + test_count, 1, value= test_count)
    # connect = "adb connect 192.168.8.107:5555"
    # subprocess.call(connect)
    time.sleep(3)
    start_modeC = "adb shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
    subprocess.call(start_modeC, shell=True, timeout=30)
    time.sleep(3)
    root = "adb wait-for-device shell setprop service.dev.mode 1"
    subprocess.call(root, shell=True, timeout=60)
    time.sleep(3)

    # 检查有无crash
    result_tomb = adb_command("adb shell ls data/tombstones/")
    result_anr = adb_command("adb shell ls data/anr/")
    if result_tomb.find("_") > 0 or result_anr.find("_") > 0:
        now_time = time.strftime("%H_%M_%S")
        subprocess.call(f"adb logcat -d > ./LOG/crash_{now_time}_{test_count}次.txt", shell=True)
        subprocess.call(f"adb pull /data/tombstones  ./tomb/{now_time}/")
        time.sleep(2)
        subprocess.call("adb shell rm -rf /data/tombstones/*", shell=True)
        subprocess.call("adb shell rm -rf /data/anr/*", shell=True)
        time.sleep(2)
        ding.send_text(time.strftime("%H:%M:%S") + f' 第 {test_count} 次导入测试后crash', is_at_all=False)
        sh.cell(1 + test_count, 4, value="YES")
    else:
        sh.cell(1 + test_count, 4, value="NO")
        print(f"第 {test_count} 次，当前无crash")

    # 检查slam初始化#
    init_command = 'adb logcat -d | findstr "delivered."'
    p_obj = subprocess.Popen(
        args=init_command,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
    if msg.find("6DOF is ready!!!!!!!!!") >= 0:
        # result = re.findall(r"6DOF is ready!!!!!!!!!")
        sh.cell(1 + test_count, 2, value="PASS")
        print("初始化成功")
    else:
        sh.cell(1 + test_count, 2, value="FAIL")
        print("初始化失败")
        time.sleep(2)


    # 检查mapping重定位
    recenter_command = 'adb logcat -d | findstr "Merge"'
    p_obj = subprocess.Popen(
        args=recenter_command,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
    time.sleep(5)
    msg = adb_command(recenter_command)
    if msg.find("mapping Merge global map") >= 0:
        sh.cell(1 + test_count, 3, value="PASS")
        print("mapping 重定位成功")
    else:
        sh.cell(1 + test_count, 3, value="FAIL")
        print("mapping 重定位失败")
        now_time = time.strftime("%H_%M_%S")
        subprocess.call(f"adb logcat -d > ./LOG/{now_time}_{test_count}次.txt", shell=True)
        time.sleep(3)
        ding.send_text(time.strftime("%H:%M:%S") + f' 第 {test_count} 次导入测试后未重定位', is_at_all=False)


    workbook1.save(Excel_File)
    # 清空地图数据
    # subprocess.call("adb shell rm -rf /sdcard/Maps/.")
    # time.sleep(2)
    # subprocess.call("adb shell rm -rf /data/misc/trackingservice/algs_bs/.")
    # time.sleep(2)
    # subprocess.call("adb shell setprop persist.yvr.large_space_enable 0")
    # print("set 0")
    # time.sleep(2)
    subprocess.call("adb reboot")
    time.sleep(85)
    print("reboot")
    test_count += 1
workbook1.save(Excel_File)


