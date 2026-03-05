# -*- coding: utf-8 -*-
import os
import subprocess
import time
import platform
import openpyxl
import shutil
import socket
import sys
import threading
from dingtalkchatbot.chatbot import DingtalkChatbot
from datetime import datetime
import win32api
from matplotlib.pyplot import savefig


class SlamRebootTester:
    def __init__(self, config):
        self.config = config
        self.workbook = None
        self.sheet = None
        self.test_count = 1
        self.running = True
        self.stop_requested = False
        self.init_excel()

        # 根据系统选择grep命令
        self.grep_cmd = 'findstr' if platform.system() == 'Windows' else 'grep'

        # 设置停止信号处理
        if platform.system() == 'Windows':
            win32api.SetConsoleCtrlHandler(self.ctrl_handler, True)
        else:
            import signal
            signal.signal(signal.SIGINT, self.signal_handler)
            signal.signal(signal.SIGTERM, self.signal_handler)

    def ctrl_handler(self, ctrl_type):
        """Windows控制台事件处理"""
        if ctrl_type in [0, 1, 2, 5]:  # CTRL_C_EVENT, CTRL_BREAK_EVENT, CTRL_CLOSE_EVENT, CTRL_SHUTDOWN_EVENT
            self.stop_requested = True
            return True
        return False

    def signal_handler(self, signum, frame):
        """Linux/Mac信号处理"""
        self.stop_requested = True

    def init_excel(self):
        """初始化Excel结果文件"""
        if not os.path.exists(self.config['excel_file']):
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Slam_Log_MSG"
            headers = ["测试次数", "初始化是否成功", "mapping重定位成功", "crash"]
            for col, header in enumerate(headers, 1):
                self.sheet.cell(1, col, value=header)
            self.workbook.save(self.config['excel_file'])
        else:
            self.workbook = openpyxl.load_workbook(self.config['excel_file'])
            self.sheet = self.workbook["Slam_Log_MSG"]
            # 查找最后一行
            self.test_count = self.sheet.max_row  # 获取最大行数，新测试从下一行开始

    def run_adb_command(self, command, capture_output=True, timeout=30):
        """执行ADB命令并返回结果"""
        full_cmd = f"adb {self.config['device_sn']} {command}"
        try:
            result = subprocess.run(
                full_cmd,
                shell=True,
                stdout=subprocess.PIPE if capture_output else None,
                stderr=subprocess.PIPE,
                text=True,
                encoding='utf-8',  # 明确指定编码
                timeout=timeout
            )
            return result.stdout.strip() if capture_output else None
        except subprocess.TimeoutExpired:
            print(f"命令超时: {full_cmd}")
            return None
        except Exception as e:
            print(f"命令执行错误: {e}")
            return None

    def check_device_connection(self):
        """检查设备连接状态"""
        output = self.run_adb_command("devices")
        return self.config['device_id'] in output if output else False

    def clear_logs(self):
        """清除设备日志"""
        self.run_adb_command("logcat -c", capture_output=False)

    def device_sleep(self):
        """使设备进入休眠"""
        self.run_adb_command("shell input keyevent 223", capture_output=False)

    def device_wake(self):
        """唤醒设备"""
        self.run_adb_command("shell input keyevent 224", capture_output=False)

    def device_reboot(self):
        """重启设备"""
        self.run_adb_command("shell reboot", capture_output=False)

    def device_root(self):
        self.run_adb_command("shell setprop service.dev.mode 1",capture_output=False)
        time.sleep(5)
        print("root")

    def check_log_for_keyword(self, keyword):
        """检查日志中是否包含关键词"""
        log_output = self.run_adb_command(f"logcat -d | {self.grep_cmd} \"{keyword}\"")
        return keyword in log_output if log_output else False

    def check_crash(self):
        """检查crash日志"""
        tomb = self.run_adb_command("shell ls /data/tombstones/")
        anr = self.run_adb_command("shell ls /data/anr/")
        # 检查是否有tombstone文件（通常文件名包含下划线）或anr文件
        tomb_has_crash = tomb and any('_' in line for line in tomb.splitlines())
        anr_has_crash = anr and any('_' in line for line in anr.splitlines())
        return tomb_has_crash or anr_has_crash

    def save_logs(self,file_name):
        """保存crash日志"""
        timestamp = datetime.now().strftime("%H_%M_%S")
        # 保存logcat
        if file_name == "crash":
            log_file = os.path.join(self.config['log_dir'], f"{file_name}_{timestamp}_{self.test_count}.txt")
            self.run_adb_command(f"logcat -d > {log_file}")
            # 拉取tombstones
            tomb_dir = os.path.join(self.config['tomb_dir'], f"crash_{timestamp}_{self.test_count}")
            os.makedirs(tomb_dir, exist_ok=True)
            self.run_adb_command(f"pull /data/tombstones/ {tomb_dir}/")
            # 清理设备tomb
            self.run_adb_command("shell rm -rf /data/tombstones/*", capture_output=False)
            self.run_adb_command("shell rm -rf /data/anr/*", capture_output=False)
        else:
            log_file = os.path.join(self.config['log_dir'], f"{file_name}_{timestamp}_{self.test_count}.txt")
            self.run_adb_command(f"logcat -d > {log_file}")

    def send_dingtalk_alert(self, message):
        """发送钉钉告警"""
        if self.config['dingtalk_webhook']:
            ding = DingtalkChatbot(self.config['dingtalk_webhook'])
            ding.send_text(f"{datetime.now().strftime('%H:%M:%S')} {message}", is_at_all=False)

    def run_test_cycle(self):
        """执行单次测试循环"""
        try:
            # 检查是否停止请求
            if self.stop_requested:
                print("停止请求已收到，正在完成当前测试...")
                return False

            # 记录测试次数
            row = self.test_count + 1  # 第一行是标题，第二行开始是第一次测试
            self.sheet.cell(row, 1, value=self.test_count)

            self.device_reboot()
            print("reboot")
            time.sleep(self.config['sleep_duration'])

            if not self.check_device_connection():
                time.sleep(5)
                self.run_adb_command(f"connect {self.config['device_id']}", capture_output=False)
                print("尝试连接设备")


            # root设备
            self.device_root()
            time.sleep(5)
            self.run_adb_command(f"connect{self.config['device_id']}", capture_output=False)

            # 检查初始化
            init_success = self.check_log_for_keyword("6DOF is ready!!!!!!")
            self.sheet.cell(row, 2, value="PASS" if init_success else "FAIL")
            print(f"初始化: {'成功' if init_success else '失败'}")
            if init_success:
                pass
            else:
                self.save_logs("slam_init")
                self.send_dingtalk_alert(f'第 {self.test_count} 次重启后初始化失败')



            # 检查重定位
            relocal_success = self.check_log_for_keyword("large_space_map_recognized setRecenterPose")
            self.sheet.cell(row, 3, value="PASS" if relocal_success else "FAIL")
            print(f"重定位: {'成功' if relocal_success else '失败'}")


            # 检查crash
            has_crash = self.check_crash()
            self.sheet.cell(row, 4, value="YES" if has_crash else "NO")
            if has_crash:
                self.save_logs("crash")
                self.send_dingtalk_alert(f'第 {self.test_count} 次重启测试后crash')
                print(f"第 {self.test_count} 次测试检测到crash")
                # return False
            else:
                print(f"第 {self.test_count} 次测试无crash")

            # 保存结果并增加计数
            self.workbook.save(self.config['excel_file'])
            self.test_count += 1
            return True

        except Exception as e:
            print(f"测试循环出错: {str(e)}")
            return False

    def archive_results(self):
        """归档测试结果到共享目录"""
        try:
            # 创建归档目录
            excel_base = os.path.splitext(os.path.basename(self.config['excel_file']))[0]
            archive_dir = os.path.join(self.config['archive_dir'], excel_base)
            os.makedirs(archive_dir, exist_ok=True)

            print(f"正在归档测试结果到: {archive_dir}")

            # 移动Excel文件
            shutil.copy(self.config['excel_file'], archive_dir)

            # 移动日志目录
            log_dest = os.path.join(archive_dir, "LOG")
            if os.path.exists(self.config['log_dir']):
                shutil.copytree(self.config['log_dir'], log_dest)

            # 移动tomb目录
            tomb_dest = os.path.join(archive_dir, "tomb")
            if os.path.exists(self.config['tomb_dir']):
                shutil.copytree(self.config['tomb_dir'], tomb_dest)

            # 上传到共享盘
            if self.config['network_share']:
                share_path = os.path.join(self.config['network_share'], excel_base)
                if os.path.exists(share_path):
                    shutil.rmtree(share_path, ignore_errors=True)
                shutil.copytree(archive_dir, share_path)
                print(f"结果已上传到共享目录: {share_path}")

            return True
        except Exception as e:
            print(f"归档结果时出错: {str(e)}")
            return False

    def continuous_testing(self):
        """连续执行测试"""
        print(f"开始SLAM重启测试，设备: {self.config['device_id']}")
        print(f"按Ctrl+C停止测试 (最多测试 {self.config['max_iterations']} 次)")

        try:
            iteration = 0
            while self.running and iteration < self.config['max_iterations']:
                # 检查停止请求
                if self.stop_requested:
                    print("停止请求已收到，正在终止测试...")
                    break

                # 检查设备连接
                if not self.check_device_connection():
                    self.run_adb_command(f"connect {self.config['device_id']}", capture_output=False)
                    time.sleep(5)
                    # 连接后再次检查
                    if not self.check_device_connection():
                        print(f"无法连接设备: {self.config['device_id']}, 等待重试...")
                        time.sleep(10)
                        continue

                # 执行测试循环
                if not self.run_test_cycle():
                    print("检测到初始化失败或crash，停止测试")
                    break

                iteration += 1
                print(f"完成第 {self.test_count - 1} 次测试 ({iteration}/{self.config['max_iterations']})")
                time.sleep(self.config['interval'])


                # 检查是否达到最大次数
                if iteration >= self.config['max_iterations']:
                    print(f"已完成最大测试次数: {self.config['max_iterations']}")
                    break

        except KeyboardInterrupt:
            print("\n用户中断测试...")
        finally:
            # 归档结果
            print("正在归档测试结果...")
            if self.archive_results():
                print("结果归档成功!")
            else:
                print("结果归档失败，请手动保存")

            # 清理资源
            if self.workbook:
                self.workbook.save(self.config['excel_file'])
                self.workbook.close()

            print("测试结束")


if __name__ == "__main__":
    # 配置参数
    config = {
        'device_sn': '-s 192.168.8.101:5555',  # 设备序列号
        'device_id': '192.168.8.101:5555',  # 设备ID
        'excel_file': './Reboot_D2_ENT_2.1.1.96_OnlyHand_0.4.66.xlsx',
        'log_dir': './LOG',
        'tomb_dir': './tomb',
        'archive_dir': './Test_Results',  # 本地归档目录
        'network_share': r'\\192.168.9.247\share\测试\aaa_cj\长跑测试记录',  # 网络共享目录
        'dingtalk_webhook': 'https://oapi.dingtalk.com/robot/send?access_token=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622',
        'sleep_duration': 60,  # 休眠时长(秒)
        'wake_delay': 5,  # 唤醒后等待时间(秒)
        'interval': 5,  # 测试间隔(秒)
        'max_iterations': 1000  # 最大测试次数 (0表示无限)
    }

    # 创建日志目录
    os.makedirs(config['log_dir'], exist_ok=True)
    os.makedirs(config['tomb_dir'], exist_ok=True)
    os.makedirs(config['archive_dir'], exist_ok=True)

    # 检查网络共享是否可达
    if config['network_share']:
        share_path = config['network_share']
        try:
            # 尝试访问共享目录
            if not os.path.exists(share_path):
                print(f"警告: 网络共享目录不可访问: {share_path}")
                # 尝试创建目录
                try:
                    os.makedirs(share_path, exist_ok=True)
                    print(f"已创建共享目录: {share_path}")
                except:
                    print("无法创建共享目录，结果将仅保存在本地")
                    config['network_share'] = None
        except:
            print("网络共享目录访问异常，结果将仅保存在本地")
            config['network_share'] = None

    tester = SlamRebootTester(config)
    tester.continuous_testing()