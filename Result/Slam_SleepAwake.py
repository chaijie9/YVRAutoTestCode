import datetime
import os
import re
import subprocess
import time
import openpyxl
import datetime as dt

Excel_File = r"./sleep_awake.xlsx"

# root 设备
# start_modeC = "adb shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
# print(start_modeC)
# subprocess.call(start_modeC, shell=True, timeout=5)
# print("start ")
# root = "adb wait-for-device shell setprop service.dev.mode 1"
# print(root)
# subprocess.call(root, shell=True, timeout=5)
# print("==============Root==============")
#
# # 关闭静置检测
# check = "adb shell getprop persist.sys.movecheck.enable"
# off_check = "adb shell setprop persist.sys.movecheck.enable 0"
# if subprocess.call(check, shell=True, timeout=15) is True:
#     subprocess.call(off_check, shell=True, timeout=15)
#
# # 修改设备休眠时间
# update_sleep_time = "adb shell setprop persist.yvr.screenofftimer 5000"
# print(update_sleep_time)
# subprocess.call(update_sleep_time, shell=True, timeout=15)
# cmd_reboot = "adb reboot "
# subprocess.call(cmd_reboot)
# time.sleep(30)
#
# # root 设备
# start_modeC = "adb shell am broadcast -a com.yvr.demo.action.dev.mode --include-stopped-packages"
# print(start_modeC)
# subprocess.call(start_modeC, shell=True, timeout=5)
# print("start ")
# root = "adb wait-for-device shell setprop service.dev.mode 1"
# print(root)
# subprocess.call(root, shell=True, timeout=5)
# print("==============Root==============")

# 当前路径创建excel保存测试结果
if not os.path.exists(Excel_File):
    wb = openpyxl.Workbook()
    wb.create_sheet(title=str("Slam_Log_MSG"), index=0)
    wb.save(Excel_File)

workbook1 = openpyxl.load_workbook(Excel_File)
sh = workbook1["Slam_Log_MSG"]
ItemIndex = ["测试次数", "Slam初始化时间", "地图切换开始时间", "地图切换完成时间", "地图切换耗时",
             "距离", "Bg", ]

count = 1
for Item in ItemIndex:
    sh.cell(1, count, value=Item)
    count = count + 1
workbook1.save(Excel_File)

test_count = 1
while test_count <= 2:

    # 清空log
    update_sleep_time = "adb shell logcat -c"
    subprocess.call(update_sleep_time, shell=True, timeout=15)
    sh.cell(1 + test_count, 1, value=test_count)

    # 灭屏,如亮屏则灭屏
    check_screen_onoff = "adb shell dumpsys deviceidle | grep mScreenOn"
    power_onoff = "adb shell input keyevent POWER"
    screen_state = subprocess.Popen(check_screen_onoff, shell=True, stdout=subprocess.PIPE)
    result = "mScreenOn=true"
    for item in screen_state.stdout.readlines():
        if str(result) in str(item):
            subprocess.call(power_onoff, shell=True, timeout=15)
        else:
            pass
        time.sleep(10)

    # 亮屏
    screen_state = subprocess.Popen(check_screen_onoff, shell=True, stdout=subprocess.PIPE)
    for item in screen_state.stdout.readlines():
        if str(result) not in str(item):
            subprocess.call(power_onoff, shell=True, timeout=15)
    time.sleep(10)

    # 检查初始化时间
    command = 'adb logcat -d | grep "EnterVRMode cost"'
    p_obj = subprocess.Popen(
        args=command,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)

    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find(r"EnterVRMode cost") >= 0:
            result_time = re.findall(r"EnterVRMode cost (.*?) s", str(msg))
            result_time = result_time[0]
            sh.cell(1 + test_count, 2, value=result_time)
        else:
            sh.cell(1 + test_count, 2, value="null")

    # 地图切换
    command1 = 'adb logcat -d | grep "kidi:NotifyMapChange"'
    p_obj = subprocess.Popen(
        args=command1,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find("18446744073709551615") >= 0:
            on_time = msg[5:18]
            data1 = dt.datetime.strptime(str(on_time).replace(" ", ""), "%H:%M:%S.%f")
            sh.cell(1 + test_count, 3, value=on_time)
        else:
            off_time = msg[5:18]
            sh.cell(1 + test_count, 4, value=off_time)

            # 计算时间差 /ms
            data2 = dt.datetime.strptime(str(off_time).replace(" ", ""), "%H:%M:%S.%f")
            result = (data2 - data1).microseconds / 1000
            sh.cell(1 + test_count, 5, value=result)

    # 判断边界是否偏移(当前头盔位置与边界中心点距离)
    command1 = 'adb logcat -d | grep "mapping ComputeDist"'
    p_obj = subprocess.Popen(
        args=command1,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find("mapping ComputeDist") >= 0:
            result = re.findall(r"ComputeDist (.*?)<", msg)
            sh.cell(1 + test_count, 6, value=result[0])
        else:
            sh.cell(1 + test_count, 6, value="未找到目标值")

    # Bg
    Bg_command = 'adb logcat -d | grep "vio init"'
    p_obj = subprocess.Popen(
        args=Bg_command,
        stdin=None, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, shell=True)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "").replace(r"\n", '')
        if msg.find("vi_estimator vio init: bg norm") >= 0:
            result = re.findall(r"bg norm: (.*)", str(msg))
            sh.cell(1 + test_count, 7, value=result[0])
        else:
            sh.cell(1 + test_count, 7, value="null")

    test_count += 1
    workbook1.save(Excel_File)
