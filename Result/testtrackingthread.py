import os
import subprocess

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

version = 'Dev_1.0.11'
sn = '1234567'
# 填写运行时间，单位为秒
duration = 10




def getpidthreads(duration, pid):
    results = subprocess.getoutput('adb shell top -d 1 -H -O PRI -O PR -O cpu -b -n ' + str(duration) + ' -p ' + pid)
    resultslist = results.splitlines()
    if os.path.exists('./' + version):
        pass
    else:
        os.mkdir('./' + version)
    fp = open('./' + version + '/' + 'threadlog.txt', 'w+')
    for line in resultslist:
        fp.write(line+'\n')
    fp.close()


def getthreadinfo(threadname):
    fp = open('./' + version + '/' + 'threadlog.txt', 'r')
    cpuidlist = []
    cpuuselist = []
    memuselist = []
    for line in fp.readlines():
        if line.find(threadname) > 0:
            linelist = [i for i in line.split(' ') if i != '']
            # print(linelist)
            cpuid = linelist[4]
            if int(cpuid) not in cpuidlist:
                # print(cpuid)
                cpuidlist.append(int(cpuid))
            cpuuse = linelist[5]
            if float(cpuuse) != 0.0:
                # print(cpuuse)
                cpuuselist.append(float(cpuuse))
            memuse = linelist[6]
            # print(memuse)
            memuselist.append(memuse)
    fp.close()
    print(cpuuselist)
    cpuidlist.sort()
    # print(cpuidlist)
    cpuidstr = ''
    for i in cpuidlist:
        cpuidstr = cpuidstr + str(i) + ' '
    cpuidstr = cpuidstr.strip()
    # print(memuselist)
    if len(cpuuselist) == 0:
        avgcpu = 0
        maxcpu = 0
    else:
        avgcpu = round(float(sum(cpuuselist)/len(cpuuselist)), 2)
        maxcpu = max(cpuuselist)
    return avgcpu, maxcpu, cpuidstr


def createsheet(title, index=0):
    wb = openpyxl.load_workbook('tracking.xlsx')
    # print(wb.active)
    ws = wb.create_sheet(title, index)
    # 创建一个填充对象
    yellowfill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bluefill = PatternFill(start_color='00B2EC', end_color='00B2EC', fill_type='solid')
    redfill = PatternFill(start_color='FA341A', end_color='FA341A', fill_type='solid')
    greenfill = PatternFill(start_color='2FE537', end_color='2FE537', fill_type='solid')
    # 设置列的宽度为 30
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20

    # 设置行的高度为30
    # for i in range(2, 10):
    #     ws.row_dimensions[i].height = 20

    datanum = 0

    row = ws.max_row
    datanum = datanum + 1

    ws['A' + str(row + 1)] = '测试场景：'
    ws.row_dimensions[row + 1].height = 40
    ws['A' + str(row + 2)] = '线程名'
    ws['A' + str(row + 2)].fill = bluefill
    ws['B' + str(row + 2)] = '单核cpu占用(均值）'
    ws['B' + str(row + 2)].fill = bluefill
    ws['C' + str(row + 2)] = '单核cpu占用（峰值）'
    ws['C' + str(row + 2)].fill = bluefill
    ws['D' + str(row + 2)] = '运行核'
    ws['D' + str(row + 2)].fill = bluefill

    # 目标值和结果栏
    ws['E' + str(row + 2)] = '目标均值'
    ws['E' + str(row + 2)].fill = bluefill
    ws['F' + str(row + 2)] = '目标峰值'
    ws['F' + str(row + 2)].fill = bluefill
    ws['G' + str(row + 2)] = '均值结果'
    ws['G' + str(row + 2)].fill = bluefill
    ws['H' + str(row + 2)] = '峰值结果'
    ws['H' + str(row + 2)].fill = bluefill
    ws.row_dimensions[row + 2].height = 30
    # 调试新遍历方法

    threadlist = ['SLAM_Image', 'SLAM_MAPPING', 'SLAM_DenseMap', 'SLAM_DepthFilte', 'SLAM_Detection', 'SLAM_ObsAvoid', 'CONTROLLER_Imag','handtrack_main','handtrack_1','handtrack_2','handtrack_3','handtrack_4']

    for i in range(len(threadlist)):
        ws.row_dimensions[row + 3 + i].height = 20
        ws['A' + str(row + 3 + i)] = threadlist[i]
        ws['A' + str(row + 3 + i)].fill = yellowfill
        ws['B' + str(row + 3 + i)] = datalist[i][0]
        ws['C' + str(row + 3 + i)] = datalist[i][1]
        ws['D' + str(row + 3 + i)] = datalist[i][2]
        # 目标值和测试结果
        # ws['E' + str(row + 3 + i)] = TARGETlist[datanum - 1][i][0]
        # ws['F' + str(row + 3 + i)] = TARGETlist[datanum - 1][i][1]
        # if DATALIST[datanum - 1][i][0] > TARGETlist[datanum - 1][i][0]:
        #     ws['G' + str(row + 3 + i)] = 'Fail'
        #     ws['G' + str(row + 3 + i)].fill = redfill
        # else:
        #     ws['G' + str(row + 3 + i)] = 'Pass'
        #     ws['G' + str(row + 3 + i)].fill = greenfill
        #
        # if DATALIST[datanum - 1][i][1] > TARGETlist[datanum - 1][i][1]:
        #     ws['H' + str(row + 3 + i)] = 'Fail'
        #     ws['H' + str(row + 3 + i)].fill = redfill
        # else:
        #     ws['H' + str(row + 3 + i)] = 'Pass'
        #     ws['H' + str(row + 3 + i)].fill = greenfill


        # ws['A' + str(row + 3)] = 'SLAM_image'
        # ws['A' + str(row + 3)].fill = yellowfill
        # ws['B' + str(row + 3)] = DATALIST[datanum - 1][0][0]
        # ws['C' + str(row + 3)] = DATALIST[datanum - 1][0][1]
        # ws['D' + str(row + 3)] = DATALIST[datanum - 1][0][2]
        # ws['A' + str(row + 4)] = 'SLAM_MAPPING'
        # ws['A' + str(row + 4)].fill = yellowfill
        # ws['B' + str(row + 4)] = DATALIST[datanum - 1][1][0]
        # ws['C' + str(row + 4)] = DATALIST[datanum - 1][1][1]
        # ws['D' + str(row + 4)] = DATALIST[datanum - 1][1][2]
        # ws['A' + str(row + 5)] = 'SLAM_DenseMap'
        # ws['A' + str(row + 5)].fill = yellowfill
        # ws['B' + str(row + 5)] = DATALIST[datanum - 1][2][0]
        # ws['C' + str(row + 5)] = DATALIST[datanum - 1][2][1]
        # ws['D' + str(row + 5)] = DATALIST[datanum - 1][2][2]
        # ws['A' + str(row + 6)] = 'SLAM_Dept hFilte'
        # ws['A' + str(row + 6)].fill = yellowfill
        # ws['B' + str(row + 6)] = DATALIST[datanum - 1][3][0]
        # ws['C' + str(row + 6)] = DATALIST[datanum - 1][3][1]
        # ws['D' + str(row + 6)] = DATALIST[datanum - 1][3][2]
        # ws['A' + str(row + 7)] = 'SLAM_Detection'
        # ws['A' + str(row + 7)].fill = yellowfill
        # ws['B' + str(row + 7)] = DATALIST[datanum - 1][4][0]
        # ws['C' + str(row + 7)] = DATALIST[datanum - 1][4][1]
        # ws['D' + str(row + 7)] = DATALIST[datanum - 1][4][2]
        # ws['A' + str(row + 8)] = 'SLAM_ObsAvoid'
        # ws['A' + str(row + 8)].fill = yellowfill
        # ws['B' + str(row + 8)] = DATALIST[datanum - 1][5][0]
        # ws['C' + str(row + 8)] = DATALIST[datanum - 1][5][1]
        # ws['D' + str(row + 8)] = DATALIST[datanum - 1][5][2]
        # 新增线程行
        # ws['A' + str(row + 9)] = 'CONTROLLER_Imag'
        # ws['A' + str(row + 9)].fill = yellowfill
        # ws['B' + str(row + 9)] = DATALIST[datanum - 1][6][0]
        # ws['C' + str(row + 9)] = DATALIST[datanum - 1][6][1]
        # ws['D' + str(row + 9)] = DATALIST[datanum - 1][6][2]
        # ws['A' + str(row + 10)] = 'handtrack_main'
        # ws['A' + str(row + 10)].fill = yellowfill
        # ws['B' + str(row + 10)] = DATALIST[datanum - 1][7][0]
        # ws['C' + str(row + 10)] = DATALIST[datanum - 1][7][1]
        # ws['D' + str(row + 10)] = DATALIST[datanum - 1][7][2]
        # ws['A' + str(row + 11)] = 'ALG_SEETHROUGH'
        # ws['A' + str(row + 11)].fill = yellowfill
        # ws['B' + str(row + 11)] = DATALIST[datanum - 1][8][0]
        # ws['C' + str(row + 11)] = DATALIST[datanum - 1][8][1]
        # ws['D' + str(row + 11)] = DATALIST[datanum - 1][8][2]


    # 设置单元格内容的水平和垂直对齐方式为居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


    wb.save('tracking.xlsx')




threadlist = ['SLAM_Image', 'SLAM_MAPPING', 'SLAM_DenseMap', 'SLAM_DepthFilte', 'SLAM_Detection', 'SLAM_ObsAvoid', 'CONTROLLER_Imag']
datalist = []

def getpid(sn, package):
    results = subprocess.getoutput('adb -s ' + sn + ' wait-for-device shell ps -A|findstr ' + package).splitlines()
    for result in results:
        if result.find('root') >= 0 or result.find('system') >= 0:
            print(result)
            pid = [i for i in result.split(' ') if i != ''][1]
            return pid
    else:
        return None

pid = getpid(sn, 'trackingservice')

getpidthreads(duration, pid)

for thread in threadlist:

    tmplist = []
    avgcpu, maxcpu, cpuidstr = getthreadinfo(thread)
    print(avgcpu)
    print(maxcpu)
    print(cpuidstr)
    tmplist.append(avgcpu)
    tmplist.append(maxcpu)
    tmplist.append(cpuidstr)
    datalist.append(tmplist)

createsheet(version)










