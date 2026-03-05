import os
import re
import subprocess
import time
import openpyxl
import pandas as pd
from dingtalkchatbot.chatbot import DingtalkChatbot

Excel_File = r"./Hand.xlsx"
if not os.path.exists(Excel_File):
    wb = openpyxl.Workbook()
    wb.create_sheet(title=str("Hand_MSG"), index=0)
    wb.save(Excel_File)

workbook1 = openpyxl.load_workbook(Excel_File)
sh = workbook1["Hand_MSG"]
ItemIndex = ["测试次数", "手柄切手势", "切换耗时", "手势切手柄", "切换耗时"]

count = 1
for Item in ItemIndex:
    sh.cell(1, count, value=Item)
    count = count + 1
workbook1.save(Excel_File)

# 测试次数
test_count = 1
while test_count <= 3:

    # 清除log
    clear_log = "adb shell logcat -c"
    print(clear_log)
    sh.cell(1 + test_count, 1, value=test_count)
    subprocess.call(clear_log, shell=True, timeout=15)

    # 断开虚拟手柄连接
    off_controller = "adb shell dumpsys android.yvr.trackingservice simulate disconnect all"
    off_controller_state = subprocess.Popen(off_controller, shell=True, stdout=subprocess.PIPE)
    time.sleep(3)

    # 校验手柄切换手势是否成功
    commond = "adb logcat -d | findstr OnInputDeviceChanged"
    p_obj = subprocess.Popen(
        args=commond, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, stdin=None, shell=True)
    time.sleep(2)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find("controller -> hand") >= 0:
            result_state = re.findall(r"input device changed:(.*)", str(msg))[0]
            sh.cell(1 + test_count, 2, value="手柄切手势成功")
        else:
            sh.cell(1 + test_count, 2, value="手柄切手势失败")

    # 记录耗时
    start_hand_commond = "adb logcat -d | findstr HandTrackingEnable"
    p_obj = subprocess.Popen(
        args=start_hand_commond, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, stdin=None, shell=True)
    time.sleep(2)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find("HandTrackingEnable") >= 0:
            start_hand_time = re.findall(r"startTracking: HandTrackingEnable cost(.*)", msg)[0]
            # print(start_hand_time)
            sh.cell(1 + test_count, 3, value=start_hand_time)
        else:
            sh.cell(1 + test_count, 3, value="null")

    # 清空log
    sh.cell(1 + test_count, 1, value=test_count)
    subprocess.call(clear_log, shell=True, timeout=15)

    # 连接虚拟手柄
    on_controller = "adb shell dumpsys android.yvr.trackingservice simulate connect all"
    on_controller_state = subprocess.Popen(on_controller, shell=True, stdout=subprocess.PIPE)
    time.sleep(2)

    # 校验手势切换手柄是否成功
    commond = "adb logcat -d | findstr OnInputDeviceChanged"
    p_obj = subprocess.Popen(
        args=commond, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, stdin=None, shell=True)
    time.sleep(2)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find("hand -> controller") >= 0:
            result_state = re.findall(r"input device changed:(.*)", str(msg))[0]
            sh.cell(1 + test_count, 4, value="手势切手柄成功")
        else:
            sh.cell(1 + test_count, 4, value="手势切手柄失败")

    # 记录耗时
    start_hand_commond = "adb logcat -d | findstr HandTrackingEnable"
    p_obj = subprocess.Popen(
        args=start_hand_commond, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, stdin=None, shell=True)
    time.sleep(2)
    for line in p_obj.stdout:
        msg = str(line).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        if msg.find("HandTrackingEnable") >= 0:
            stop_hand_time = re.findall(r"stopTracking: HandTrackingEnable cost(.*)", msg)[0]
            # print(start_hand_time)
            sh.cell(1 + test_count, 5, value=stop_hand_time)
        else:
            sh.cell(1 + test_count, 5, value="null")

        test_count += 1
    workbook1.save(Excel_File)


df = pd.read_excel(r'./Hand.xlsx', sheet_name='Hand_MSG')
data = df.values
str_data = str(data)
webhook = 'https://oapi.dingtalk.com/robot/send?access_token=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622'
ding = DingtalkChatbot(webhook)
ding.send_text(msg="手柄/手势切换测试结果\n"
                   "\t手柄切手势\t切换耗时\t\t      手势且手柄\t\t切换耗时\n"
                   "" + str_data.replace("[", '').replace("]", ''), is_at_all=False)


'''

'''