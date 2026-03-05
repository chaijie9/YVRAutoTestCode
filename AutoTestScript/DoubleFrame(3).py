# rem adb shell dumpsys vrruntimeservice_native -c 0x00100000
# adb shell dumpsys vrruntimeservice_native -c 00200000
# timeout /t 300
# adb shell dumpsys vrruntimeservice_native -c 0
#
# timeout /t 3
# adb pull sdcard/misc/yvr/accuracy_test

import os
import re
import shutil
import subprocess
import time
from dataclasses import dataclass
from decimal import Decimal
from typing import Iterable, Optional

import matplotlib.pyplot as plt
import openpyxl


@dataclass(frozen=True)
class Config:
    check_time: int = 300
    excel_file: str = "double_frame.xlsx"
    log_file: str = "./fps.txt"
    repeat_frame_file: str = "./accuracy_test/repeat_frame.txt"
    pull_dir: str = "./accuracy_test"
    out_dir: str = "./test_data"


HEADERS = [
    "场景",
    "DoubleFrame",
    "最大渲染次数",
    "max_FPS",
    "min_FPS",
    "avg_FPS",
    "max_GPU",
    "min_GPU",
    "avg_GPU",
    "max_CPU",
    "min_CPU",
    "avg_CPU",
]


def adb_call(cmd: str) -> int:
    return subprocess.call(cmd, shell=True)


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def init_excel(excel_file: str, sheet_name: str, headers: Iterable[str]):
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        wb.create_sheet(title=sheet_name, index=0)
        wb.save(excel_file)
    wb = openpyxl.load_workbook(excel_file)
    sh = wb[sheet_name]
    for col, header in enumerate(headers, 1):
        sh.cell(1, col, value=header)
    wb.save(excel_file)
    return wb, sh


def capture_logcat_to_file(log_path: str, duration_sec: int) -> None:
    adb_call("adb logcat -c")
    time.sleep(3)
    with open(log_path, "w", encoding="utf-8", errors="ignore") as f:
        proc = subprocess.Popen("adb logcat", stdout=f, stderr=subprocess.PIPE, shell=True)
        try:
            time.sleep(duration_sec)
        finally:
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except Exception:
                pass


def pull_accuracy_test() -> None:
    adb_call("adb pull sdcard/misc/yvr/accuracy_test")
    time.sleep(5)
    adb_call("adb kill-server")


def enable_double_frame_capture(enable: bool) -> None:
    adb_call(
        "adb shell dumpsys vrruntimeservice_native -c 00200000"
        if enable
        else "adb shell dumpsys vrruntimeservice_native -c 0"
    )


def read_repeat_counts(repeat_frame_file: str) -> tuple[int, int]:
    double_frame_count = 0
    repeat_count_list: list[int] = []
    if not os.path.exists(repeat_frame_file):
        return 0, 0

    with open(repeat_frame_file, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if not line.strip():
                continue
            double_frame_count += 1
            if "frame_index" in line:
                m = re.findall(r"repeat_count:(.*)", line)
                if m:
                    try:
                        repeat_count_list.append(int(m[0].strip()))
                    except ValueError:
                        pass
    max_repeat = max(repeat_count_list) if repeat_count_list else 0
    return double_frame_count, max_repeat


def normalize_log_file_inplace(path: str) -> None:
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = [line for line in f if line.strip()]
    if lines:
        lines.pop()  # 删除最后一行（通常是不完整的收尾行）
    with open(path, "w", encoding="utf-8") as f:
        f.writelines(lines)


def parse_metrics_from_log(path: str):
    fps_list: list[float] = []
    gpu_list: list[float] = []
    cpu_list: list[float] = []
    miss_l = 0
    miss_r = 0

    if not os.path.exists(path):
        return fps_list, gpu_list, cpu_list, miss_l, miss_r

    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if "FPS=" in line:
                m = re.findall(r"FPS=(.*?)/", line)
                if m:
                    try:
                        fps_list.append(float(m[0]))
                    except ValueError:
                        pass

            if ", GPU%=" in line:
                m = re.findall(r", GPU%=(.*)", line)
                if m:
                    try:
                        gpu_list.append(float(m[0]))
                    except ValueError:
                        pass

            if ", CPU%=" in line:
                m = re.findall(r", CPU%=(.*),", line)
                if m:
                    try:
                        cpu_list.append(float(m[0]))
                    except ValueError:
                        pass
            elif "E DataMonitor: cam_rvst frame miss" in line:
                miss_r += 1
            elif "E DataMonitor: cam_lvst frame miss" in line:
                miss_l += 1

    return fps_list, gpu_list, cpu_list, miss_l, miss_r


def stats(values: list[float]) -> tuple[float, float, Decimal]:
    if not values:
        return 0.0, 0.0, Decimal("0.00")
    max_v = max(values)
    min_v = min(values)
    avg_v = Decimal(sum(values) / len(values)).quantize(Decimal("0.00"))
    return max_v, min_v, avg_v


def save_plots(out_dir: str, tag: str, cpu_list: list[float], gpu_list: list[float], fps_list: list[float]) -> None:
    ensure_dir(out_dir)

    fig, ax = plt.subplots(1, 2)
    ax[0].plot(cpu_list)
    ax[0].set_ylabel("cpu")
    ax[1].plot(gpu_list)
    ax[1].set_ylabel("gpu")
    ax[0].set_xlim()
    ax[0].set_ylim(0, 1)
    ax[1].set_xlim()
    ax[1].set_ylim(0, 1)
    plt.savefig(os.path.join(out_dir, f"{tag}CPU&GPU.jpg"))
    plt.show()

    plt.hlines(89, 0, 300, colors="green", linestyles="--")
    plt.xlabel("")
    plt.ylabel("FPS")
    plt.xlim(0, 300)
    plt.ylim(20, 105)
    plt.plot(fps_list)
    plt.savefig(os.path.join(out_dir, f"{tag}FPS.jpg"))
    plt.show()


def main(cfg: Optional[Config] = None) -> None:
    cfg = cfg or Config()
    now_tag = time.strftime("%Y%m%d%H%M%S", time.localtime())
    test_count = 1

    time.sleep(3)
    capture_logcat_to_file(cfg.log_file, cfg.check_time)

    enable_double_frame_capture(True)
    time.sleep(cfg.check_time)
    enable_double_frame_capture(False)

    time.sleep(5)
    adb_call("adb shell input keyevent POWER")
    pull_accuracy_test()

    workbook1, sh = init_excel(cfg.excel_file, "double_frame", HEADERS)

    double_frame_count, max_repeat = read_repeat_counts(cfg.repeat_frame_file)
    row = 1 + test_count
    sh.cell(row, 2, value=double_frame_count)
    sh.cell(row, 3, value=max_repeat)
    workbook1.save(cfg.excel_file)

    print("=======================")
    print("DoubleFrame次数：", double_frame_count)
    print("最大渲染次数", max_repeat)

    normalize_log_file_inplace(cfg.log_file)
    ensure_dir(cfg.out_dir)
    shutil.copy2(cfg.log_file, os.path.join(cfg.out_dir, f"{now_tag}.txt"))
    if os.path.exists(cfg.repeat_frame_file):
        shutil.copy2(cfg.repeat_frame_file, os.path.join(cfg.out_dir, f"repeat{now_tag}.txt"))

    fps_list, gpu_list, cpu_list, miss_l, miss_r = parse_metrics_from_log(cfg.log_file)
    print(miss_l, miss_r)

    max_fps, min_fps, avg_fps = stats(fps_list)
    max_gpu, min_gpu, avg_gpu = stats(gpu_list)
    max_cpu, min_cpu, avg_cpu = stats(cpu_list)

    print("=======================")
    print("FPS最大值： ", max_fps)
    print("FPS最小值： ", min_fps)
    print("FPS平均值： ", avg_fps)
    print("=======================")
    print("GPU最大值： ", max_gpu)
    print("GPU最小值： ", min_gpu)
    print("GPU平均值： ", avg_gpu)
    print("=======================")
    print("CPU最大值：", max_cpu)
    print("CPU最小值：", min_cpu)
    print("CPU平均值：", avg_cpu)

    save_plots(cfg.out_dir, now_tag, cpu_list, gpu_list, fps_list)


if __name__ == "__main__":
    main()










