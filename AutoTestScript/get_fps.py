# -*- coding: utf-8 -*-
# rem adb shell dumpsys vrruntimeservice_native -c 0x00100000
# adb shell dumpsys vrruntimeservice_native -c 00200000
# timeout /t 300
# adb shell dumpsys vrruntimeservice_native -c 0
#
# timeout /t 3
# adb pull sdcard/misc/yvr/accuracy_test

#
import os
import re
import shutil
import subprocess
import time
from decimal import Decimal

import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

check_time = 10

Excel_File = r"double_frame.xlsx"
filename = r"./fps.txt"
# filename = r"./test_data/20250731153808.txt"
# source_repeat = "./accuracy_test/repeat_frame.txt"
Now_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
FpsCount = 0
GpuCount = 0
CpuCount = 0
l = 0
r = 0
FPS_list = []
GPU_list = []
Cpu_list = []
double_frame_count = 0
repeat_count = 0
repeat_count_list = []
test_count = 1
#
# time.sleep(3)
# 抓取特定时间的log
logcmd = "adb logcat"
subprocess.call("adb logcat -c", shell=True)
logcat_file = open(filename, 'w')
time.sleep(3)
Poplog = subprocess.Popen(logcmd, stdout=logcat_file, stderr=subprocess.PIPE, shell=True)
# 开启抓取double frame开关
# subprocess.call("adb shell dumpsys vrruntimeservice_native -c 00200000", shell=True)
print("start")
time.sleep(check_time)
Poplog.terminate()
logcat_file.close()
# 停止
# subprocess.call("adb shell dumpsys vrruntimeservice_native -c 0", shell=True)
# time.sleep(5)
subprocess.call("adb shell input keyevent 223")
# 拉取
# subprocess.call("adb pull sdcard/misc/yvr/accuracy_test", shell=True)
time.sleep(3)
subprocess.call("adb kill-server", shell=True)



# 当前路径创建excel保存测试结果
# if not os.path.exists(Excel_File):
#     wb = openpyxl.Workbook()
#     wb.create_sheet(title=str("double_frame"), index=0)
#     wb.save(Excel_File)
#
# workbook1 = openpyxl.load_workbook(Excel_File)
# sh = workbook1["double_frame"]
# ItemIndex = ["场景", "DoubleFrame", "最大渲染次数", "max_FPS", "min_FPS", "avg_FPS", "max_GPU", "min_GPU", "avg_GPU", "max_CPU", "min_CPU", "avg_CPU" ]
#
# count = 1
# for Item in ItemIndex:
#     sh.cell(1, count, value=Item)
#     count = count + 1
# workbook1.save(Excel_File)

# with open(f"accuracy_test/repeat_frame.txt") as f:
#     result = f.readlines()
#     for n in result:
#         double_frame_count += 1
#         if n.find("frame_index") >= 0:
#             repeat_count = int(re.findall('repeat_count:(.*)', n)[0])
#             repeat_count_list.append(repeat_count)
#
#  # max(repeat_count) = 最大渲染次数
# if len(repeat_count_list) != 0:
#     sh.cell(1 + test_count, 3, value=max(repeat_count_list))
#     sh.cell(1 + test_count, 2, value=double_frame_count)
# else:
#     sh.cell(1 + test_count, 3, value=0)
#     sh.cell(1 + test_count, 2, value=0)
#
# print("=======================")
# print("DoubleFrame次数：", double_frame_count)
# print("最大渲染次数", max(repeat_count_list))

# 处理源文件，删除最后一行内容
# with open(filename, 'r', encoding='utf-8') as f:  # 使用GBK编码读取
#     lines = [line for line in f if line.strip() != '']
#     lines.pop()  # 删除最后一行

# with open(filename, 'w',encoding="utf-8") as f:
#     f.writelines(lines)

# 保存单次测试数据
if os.path.exists("test_data"):
    pass

else:
    os.mkdir("test_data")
shutil.copy2(filename, f"./test_data/{Now_time}.txt")
# shutil.copy2(source_repeat, f"./test_data/repeat{Now_time}.txt")

filename = "./fps.txt"
with open(filename, "r", encoding="utf-8", errors="ignore") as f:
    result_data = f.readlines()
    for eachline in result_data:
        if eachline.find("FPS=") >= 0:
            FpsCount += 1
            fps = float(re.findall("FPS=(.*?)/", eachline)[0])
            # print(fps)
            FPS_list.append(fps)

        if eachline.find(", GPU%=") >= 0:
            GpuCount += 1
            gpu = float(re.findall(', GPU%=(.*),', eachline)[0])
            GPU_list.append(gpu)
            # print("GPU=", GPU_list)

        if eachline.find(', CPU%=') >= 0:
            CpuCount += 1
            cpu = float(re.findall(', CPU%=(.*), ', eachline)[0])
            Cpu_list.append(cpu)
            # print("CPU=", Cpu_list)
        elif eachline.find("E DataMonitor: cam_rvst frame miss") >= 0:
            r += 1
        elif eachline.find("E DataMonitor: cam_lvst frame miss") >= 0:
            l += 1
    print(l, r)

# 计算max,min,avg 并写入
max_FPS = max(FPS_list)
min_FPS = min(FPS_list)
avg_FPS = sum(FPS_list) / len(FPS_list)

# print(GPU_list, "+++++++++++++++++")
max_GPU = max(GPU_list)
min_GPU = min(GPU_list)
avg_GPU = sum(GPU_list) / len(GPU_list)

max_CPU = max(Cpu_list)
min_CPU = min(Cpu_list)
avg_CPU = sum(Cpu_list) / len(Cpu_list)

avg_FPS = Decimal(avg_FPS).quantize(Decimal("0.00"))
avg_GPU = Decimal(avg_GPU).quantize(Decimal("0.00"))
avg_CPU = Decimal(avg_CPU).quantize(Decimal("0.00"))


# sh.cell(1 + test_count, 4, value=max_FPS)
# sh.cell(1 + test_count, 5, value=min_FPS)
# sh.cell(1 + test_count, 6, value=avg_FPS)
# sh.cell(1 + test_count, 7, value=max_GPU)
# sh.cell(1 + test_count, 8, value=min_GPU)
# sh.cell(1 + test_count, 9, value=avg_GPU)
# sh.cell(1 + test_count, 10, value=max_CPU)
# sh.cell(1 + test_count, 11, value=min_CPU)
# sh.cell(1 + test_count, 12, value=avg_CPU)
#


print("=======================")
print('FPS最大值： ', max_FPS)
print('FPS最小值： ', min_FPS)
print('FPS平均值： ', avg_FPS)
print("=======================")
print('GPU最大值： ', max_GPU)
print('GPU最小值： ', min_GPU)
print('GPU平均值： ', avg_GPU)
print("=======================")
print('CPU最大值：', max_CPU)
print('CPU最小值：', min_CPU)
print('CPU平均值：', avg_CPU)


# # 绘图
fig, ax = plt.subplots(1, 2)
ax[0].plot(Cpu_list)
ax[0].set_ylabel("cpu")
ax[1].plot(GPU_list)
ax[1].set_ylabel("gpu")
ax[0].set_xlim()
ax[0].set_ylim(0, 1)
ax[1].set_xlim()
ax[1].set_ylim(0, 1)
plt.savefig(f"./test_data/{Now_time}CPU&GPU.jpg")
# plt.savefig(f"./test_data/{Now_time}CPU&GPU.jpg")

plt.show()
plt.hlines(89, 0, 300, colors='green', linestyles='--')
plt.xlabel('')
plt.ylabel('FPS')
plt.xlim(0, check_time)
plt.ylim(20, 105)
plt.plot(FPS_list)
# plt.savefig(f"./test_data/{Now_time}FPS.jpg")
plt.savefig(f"./test_data/{Now_time}FPS.jpg")

plt.show()









