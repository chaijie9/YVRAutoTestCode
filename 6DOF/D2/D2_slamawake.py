'''
result = subprocess.run("args", capture_output=True, text=True)
print(result.stdout)

# args：以为字符串或字符串列表
# capture_output：为 True 时可以捕获输出结果，包括标准输出和标准错误，结果保存在 stdout 和 stderr 中
# text：为 True 时以文本格式处理输入输出，False 时为字节格式
# input：用于传递输入给命令的数据，可以为字符串或字节
# timeout：最大执行时间，超时后将终止命令
# check 设置为 True 时，如果返回非零状态码，则抛出 CalledProcessError
'''
import copy
import os
import re
import subprocess
import datetime as dt
import time
import eventlet as eventlet
import openpyxl
from dingtalkchatbot.chatbot import DingtalkChatbot

webhook = 'https://oapi.dingtalk.com/robot/send?access_token' \
          '=ac6710a1ccc6f62d22e0cd8aa690986f5cf3afb2064f60e63c7a4b832d64b622'
ding = DingtalkChatbot(webhook)


def calculate_runtime(func):
    def wrapper(*args, **kwargs):
        # 记录开始时间戳
        start_time = time.time()
        # 调用原始函数
        result = func(*args, **kwargs)
        # 记录结束时间戳
        end_time = time.time()
        # 计算运行耗时
        global runtime
        runtime = end_time - start_time
        print(f"Function '{func.__name__}' runtime: {runtime:.4f} seconds")

        return result

    return wrapper


def adb_command(command):
    try:
        result = subprocess.run(command, capture_output=True, check=True, text=True, timeout=10)
        msg = str(result.stdout).replace("b'", "").replace(r"\r\n'", "").replace("'", "") \
            .replace("[", "").replace("]", "").replace(r"\t", "")
        return msg
    except subprocess.CalledProcessError as e:
        ding.send_text(msg=time.strftime("%H:%M:%S ") + f' {e} 无结果', is_at_all=False)
        return f"未获取到当前结果 {e}"
    except subprocess.TimeoutExpired as e:
        return f"{command} 命令运行超时\n{e}"


def screen_onoff():
    check_command = adb_command("adb shell dumpsys deviceidle | grp mScreenOn")
    result = check_command.strip().replace("\\n", '')
    if "mScreenOn=false" in result:
        adb_command("adb shell input keyevent POWER")
        time.sleep(15)
        print("亮屏")
    else:
        adb_command("adb shell input keyevent POWER")
        print("当前亮屏，先灭屏")
        time.sleep(15)
        adb_command("adb shell input keyevent POWER")
        print("亮屏")


# EnterVRMode cost
def enter_vrmode():
    try:
        enter = adb_command("adb shell logcat -d | grep 'EnterVRMode cost' |tail -1")
        result_time = re.findall(r"EnterVRMode cost (.*?) s", str(enter))
        print(result_time[0])
        return result_time[0]
    except IndexError as e:
        return print("未找到EnterVRMode cost")
    except Exception as e:
        return print(f"未知错误{e}")


# mapping handoff
def mapping_handoff():
    msg = adb_command('adb shell logcat -d | grep "kidi:NotifyMapChange"|tail -2')
    for line in msg.splitlines():
        if line.find("18446744073709551615") >= 0:
            on_time = line[5:18]
        else:
            off_time = line[5:18]
    # print(on_time)
    data1 = dt.datetime.strptime(str(on_time).replace(" ", ""), "%H:%M:%S.%f")
    data2 = dt.datetime.strptime(str(off_time).replace(" ", ""), "%H:%M:%S.%f")
    result = (data2 - data1).microseconds / 1000
    return on_time, off_time, result

@calculate_runtime
# slam init Bg
def slam_init():
    msg = adb_command("adb shell logcat -d | grep 'vio init'| tail -1")
    try:
        result = re.findall(r"bg norm: (.*),", str(msg))[0]
        return result
    except IndexError as e:
        return f"vio init 未找到{e}"


def boundary_offset():
    msg = adb_command("adb shell logcat -d | grep 'mapping ComputeDist'")
    result = re.findall(r"ComputeDist (.*?)<", msg)
    print(result[0])
    return result[0]

    # try:
    #     with eventlet.Timeout(1, True):
    #         msg = adb_command("adb shell logcat -d | grep 'vio init'")
    #     return msg, print("未超时")
    # except eventlet.timeout.Timeout:
    #     print("已超时")


def check_tomb(test_count):
    result_tomb = adb_command("adb shell ls data/tombstones/")
    result_anr = adb_command("adb shell ls data/anr/")
    print(result_tomb, result_anr)
    if result_tomb.find("_") > 0 or result_anr.find("_") > 0:
        ding.send_text(time.strftime("%H:%M:%S") + f' 第 {test_count} 次休眠唤醒后crash', is_at_all=False)
    else:

        print(f"第 {test_count} 次，当前无crash")


def slam_sleep_awake(excel_file, test_count):
    Excel_File = rf"./{excel_file}"
    if not os.path.exists(Excel_File):
        wb = openpyxl.Workbook()
        wb.create_sheet(title=str("Slam_Log_MSG"), index=0)
        wb.save(Excel_File)
    workbook1 = openpyxl.load_workbook(Excel_File)
    sh = workbook1["Slam_Log_MSG"]
    ItemIndex = ["测试次数", "Slam初始化时间", "地图切换开始时间", "地图切换完成时间", "地图切换耗时",
                 "距离", "Bg"]
    count = 1
    for Item in ItemIndex:
        sh.cell(1, count, value=Item)
        count = count + 1
    workbook1.save(Excel_File)

    while test_count <= 500:
        trace = 0
        screen_onoff()
        # print(slam_init())
        # init = slam_init()
        # result = "未找到"
        # if result in init:
        #     trace += 1
        # if trace != 0:
        #     ding.send_text(time.strftime("%H:%M:%S") + f' 第 {test_count} 次初始化未成功', is_at_all=False)
        #     print("STOP！！！！！！！！！！")
        #     break
        # enter_time = enter_vrmode()
        # mapping_call = mapping_handoff()
        # offset = boundary_offset()
        # Bg = slam_init()
        sh.cell(1 + test_count, 1, value=test_count)
        # sh.cell(1 + test_count, 2, value=enter_time)
        # sh.cell(1 + test_count, 3, value=mapping_call[0])
        # sh.cell(1 + test_count, 4, value=mapping_call[1])
        # sh.cell(1 + test_count, 5, value=mapping_call[2])
        # sh.cell(1 + test_count, 6, value=offset)
        # sh.cell(1 + test_count, 6, value=Bg)
        workbook1.save(Excel_File)
        check_tomb(test_count)
        test_count += 1
        print(f"当前为第{test_count}次测试")


if __name__ == '__main__':
    # slam_sleep_awake("sleep_0802.xlsx", 1)

    test_count = 1
    while test_count <= 1000:
        adb_command("adb shell input keyevent POWER")
        time.sleep(15)
        check_tomb(test_count)
        test_count += 1
        print(test_count)


# #
# print(mapping_handoff())
# def example_function(delay):
#     time.sleep(delay)
#     print(f"Function is done with delay {delay} seconds")


# s = SLAM_Sleep_Awake.adb_command("adb shell logcat -d | grep 'EnterVRMode cost'")

# print(s)
# if s is None:
#     print("none")
# else:
#     print("111")
# if __name__ == '__main__':
#     # demo_str = "demo".encode("gbk")
#     # demo = demo_str.decode("gbk").encode("utf8")
#     # # ['info', 'xiaoZhang', '33', 'shandong']
#     # s = "info：xiaoZhang 33 shandong"
#     # s = re.compile(r'\W').split(s)
#
#     question_list = [[1, 2],[3, 4],[5, 6]]
#     # result = [a for inside in question_list for a in inside]
#     # print(result)
#     res_dict = {"A": 1, "B": 2}
#     result_dic = {v: k for k, v in res_dict.items()}
#     # print(result_dic)

# a = [1, 2, [3, 4]]
# b = a
# c = a[:]
# d = a.copy()
# e = copy.deepcopy(a)
# # print(id(a), id(b), id(c), id(d), id(e))
# # a.append(5)
# # print(a, b, c, d, e)
#
# a[2][1] = 7
# # print(a, b, c, d, e)
#
# original_list = [1, 2, [3, 4]]
# deep_copy = copy.copy(original_list)
#
# print(original_list)    # [1, 2, [3, 4]]
# print(deep_copy)        # [1, 2, [3, 4]]
#
# deep_copy[2][0] = 99
# print(original_list)    # [1, 2, [3, 4]]  原始对象不受影响
# print(deep_copy)        # [1, 2, [99, 4]]